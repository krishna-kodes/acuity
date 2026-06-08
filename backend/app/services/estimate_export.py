"""Streaming export builders for the effort estimate (Phase 5)."""

import csv
import io
import json

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.sync import Epic, Task

_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
_BODY_FONT = Font(name="Arial")


def _first_assignee(raw: str | None) -> str:
    """Return the first element of a JSON-encoded assignees list, or ''."""
    if not raw:
        return ""
    try:
        parsed = json.loads(raw)
        return parsed[0] if parsed else ""
    except (json.JSONDecodeError, IndexError, TypeError):
        return ""


def _style_header(ws, headers: list[str], widths: list[int]) -> None:
    """Bold blue header row, freeze pane, autofilter."""
    for col_idx, width in enumerate(widths, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def build_estimate_csv(project_id: int, db: Session) -> io.BytesIO:
    """Return a BytesIO of UTF-8 CSV with one row per task."""
    epics = db.query(Epic).filter(Epic.project_id == project_id).all()

    strio = io.StringIO()
    writer = csv.writer(strio)
    writer.writerow([
        "epic_title", "epic_estimated_points", "task_title", "task_assignee",
        "task_estimated_points", "confidence_low", "confidence_mid", "confidence_high",
        "github_issue_url", "sync_status",
    ])

    for epic in epics:
        tasks = db.query(Task).filter(Task.epic_id == epic.id).all()
        if not tasks:
            writer.writerow([
                epic.title, epic.estimated_points or "",
                "", "", "", "", "", "", "", "",
            ])
            continue
        for task in tasks:
            pts = task.estimated_points or 0
            writer.writerow([
                epic.title,
                epic.estimated_points or "",
                task.title,
                _first_assignee(task.assignees),
                pts,
                round(pts * 0.8, 1),
                pts,
                round(pts * 1.3, 1),
                task.github_issue_url or "",
                task.sync_status.value if task.sync_status else "",
            ])

    return io.BytesIO(strio.getvalue().encode("utf-8"))


def build_estimate_xlsx(project_id: int, db: Session) -> io.BytesIO:
    """Return a BytesIO of an XLSX workbook with Summary and Task Breakdown sheets."""
    epics = db.query(Epic).filter(Epic.project_id == project_id).all()

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    headers1 = ["Epic", "Estimated Points", "Actual Points", "Variance", "% Variance", "Sync Status"]
    widths1 = [40, 18, 18, 12, 14, 14]
    ws1.append(headers1)

    for row_idx, epic in enumerate(epics, start=2):
        ws1.cell(row=row_idx, column=1, value=epic.title).font = _BODY_FONT
        ws1.cell(row=row_idx, column=2, value=epic.estimated_points).font = _BODY_FONT
        ws1.cell(row=row_idx, column=3, value=None).font = _BODY_FONT          # actual_points — not in model
        ws1.cell(row=row_idx, column=4, value=f"=C{row_idx}-B{row_idx}").font = _BODY_FONT
        ws1.cell(row=row_idx, column=5, value=f'=IF(B{row_idx}=0,"—",D{row_idx}/B{row_idx})').font = _BODY_FONT
        ws1.cell(row=row_idx, column=6, value=epic.sync_status.value if epic.sync_status else "").font = _BODY_FONT

    _style_header(ws1, headers1, widths1)

    # ── Sheet 2: Task Breakdown ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Task Breakdown")
    headers2 = ["Epic", "Task", "Assignee", "Est. Points", "Low", "Mid", "High", "GitHub Issue", "Sync Status"]
    widths2 = [35, 40, 20, 14, 8, 8, 8, 40, 14]
    ws2.append(headers2)

    task_row = 2
    for epic in epics:
        tasks = db.query(Task).filter(Task.epic_id == epic.id).all()
        for task in tasks:
            pts = task.estimated_points or 0
            row_vals = [
                epic.title,
                task.title,
                _first_assignee(task.assignees),
                pts,
                round(pts * 0.8, 1),
                pts,
                round(pts * 1.3, 1),
                task.github_issue_url or "",
                task.sync_status.value if task.sync_status else "",
            ]
            for col_idx, val in enumerate(row_vals, start=1):
                ws2.cell(row=task_row, column=col_idx, value=val).font = _BODY_FONT
            task_row += 1

    _style_header(ws2, headers2, widths2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
